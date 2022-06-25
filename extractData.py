from openpyxl import load_workbook
import re

wb = load_workbook("temp.xlsx")   # open an Excel file and return a workbook
univName = "Miami University" 
courseCodeString = "CSCI"  
if univName  in wb.sheetnames:
    print(univName+' exists')

else:
    wb.create_sheet(univName)

activesheet = wb[univName]

if(activesheet["A1"]!="Course No"):
    #activesheet.append(("Course No","Course Name"))
    activesheet["A1"]="Course No"
    activesheet["B1"]="Course Name"

prevline=""
courseCode=[]
duplicateCount=0
tempcode=""
tempcode=""

with open('temp.txt',encoding="utf8") as f:
    lines = f.readlines()
    for i in range(0, len(lines)):
        line = lines[i]   
        if("CSE" in line[:5]):
                
                #if("EN." in line[:4]):
                ##hipposi = tmpstr.find(".")+7
                tempcode = line[:8]
                
                tempName = line[8:]
                if("(" in tempName[:2]):
                    tempName = tempName[tempName.find(")")+1:]
                #tempName = tempName[:tempName.find(".")]
                #print("before tempcod "+tempcode)
                
                """
                if("(" in tempName):
                    
                    tempName = tempName[tempName.find(")")+1:]
                    #print("AFTER tempname "+tempName)
                first_digit = re.search('\d', tempName)
               # print(first_digit.start())
                tempName = tempName[:first_digit.start()]
                #elif("CSCI-C" in line[:7]):
                    tmpstr = line[5:]
                    hipposi = tmpstr.find("-")+5
                    tempcode = "CSCI-"+line[5:hipposi]
                    tempName = line[hipposi+1:]
                #else:
                    tempcode = line[:line.find("-")]
                    tempName = line[line.find("-")+1:]
                    """
                if(tempcode  in courseCode):
                    print("duplicate for : "+line+" total duplicate "+str(duplicateCount))
                    duplicateCount+=1
                    continue
                if(tempName[0]==" "):
                    tempName = tempName[1:]
                courseCode.append(tempcode)
                activesheet.append((tempcode,tempName))

            # break
                print(tempcode+" "+tempName)
        prevline = line
        
f.close()

wb.save('temp.xlsx')
wb.close()